# excel_handler.py
import pandas as pd
from openpyxl import load_workbook
from config import arquivo_excel_path
from openpyxl.styles import PatternFill

def ler_excel():
    try:
        df = pd.read_excel(arquivo_excel_path)
        return df
    except Exception as e:
        print(f"Erro ao ler o arquivo Excel: {e}")
        return pd.DataFrame()

def obter_nome_cliente(df, index):
    try:
        nome_completo = str(df.iloc[index, 0]).strip()  # Coluna 'A'
        primeiro_nome = nome_completo.split()[0]  # Extrai apenas o primeiro nome
        return primeiro_nome
    except Exception as e:
        print(f"Erro ao obter o nome do cliente: {e}")
        return ""

def obter_telefone(df, index):
    try:
        telefone = str(df.iloc[index, 1])  # Coluna 'B'
        return telefone
    except Exception as e:
        print(f"Erro ao obter o telefone: {e}")
        return ""

def colorir_linha_verdadeira(arquivo, linha, cor_hex):
    try:
        wb = load_workbook(arquivo)
        ws = wb.active
        fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
        for cell in ws[linha]:
            cell.fill = fill
        wb.save(arquivo)
    except Exception as e:
        print(f"Erro ao colorir a linha {linha}: {e}")
def obter_nome_vaga(df, index):
    try:
        nome_vaga = df.iloc[index, 2]  # Supondo que a vaga está na coluna C
        return str(nome_vaga).strip()
    except Exception as e:
        print(f"Erro ao obter o nome da vaga: {e}")
        return ""
def obter_cliente(df, index):
    try:
        nome_s = str(df.iloc[index, 3]).strip()  # Coluna 'A'
        
        return nome_s
    except Exception as e:
        print(f"Erro ao obter o nome do cliente: {e}")
        return ""
    
def obter_email(df, index):
    try:
        email = df.iloc[index, 5]  # Substitua 1 pelo índice correto da coluna de e-mail
        return str(email).strip()
    except Exception as e:
        print(f"Erro ao obter o e-mail: {e}")
        return ""
def obter_origem(df, index):
    try:
        org = df.iloc[index, 4]  # Substitua 1 pelo índice correto da coluna de e-mail
        return str(org)
    except Exception as e:
        print(f"Erro ao obter o e-mail: {e}")
        return ""